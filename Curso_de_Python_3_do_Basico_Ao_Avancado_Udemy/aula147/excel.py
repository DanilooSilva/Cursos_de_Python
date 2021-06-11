import openpyxl
from random import uniform

# pedidos = openpyxl.load_workbook(r'aula147\pedidos.xlsx')
# nomes_planilhas = pedidos.sheetnames
# planilha1 = pedidos['Planilha1']

# for linha in planilha1:
#     print(linha[0].value, end=" ")
#     print(linha[1].value, end=" ")
#     print(linha[2].value)

planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

for linha in range(1, 11):
    numero_pedidos = linha - 1
    planilha1.cell(linha, 1).value = numero_pedidos
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

for linha in range(1, 11):
    planilha2.cell(linha, 1).value = f'Danilo {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 2).value = f'Maria {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 3).value = f'Ohara {linha} {round(uniform(10, 100), 2)}'

planilha.save(r'aula147\nova_planilha.xlsx')